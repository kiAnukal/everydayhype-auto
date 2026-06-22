"""Telegram ingest agent — TRAIN THE WRITER BY SENDING THE BOT ANYTHING:
  • IMAGES (screenshots of top posts)  → vision reads them
  • TEXT (paste a post / an idea)       → LLM shapes it
  • FILES (txt / md / csv / json / pdf — e.g. Apify exports) → extracted & learned
Each is distilled into a reusable WINNING PATTERN (topic/hook/headline/why) and appended to
examples.md, which the writer studies. No typing required. Runs on a cron in GitHub Actions
(laptop off). Tracks a Telegram offset so nothing is read twice."""
import json, base64, io, requests
from openai import OpenAI
import config as C

OFFSET_FILE = C.STATE / "tg_offset.txt"
EXAMPLES = C.ROOT / "examples.md"

SYS = ("You curate a content-style library for the Instagram brand @everydayhypehq. From the given "
       "post(s) / notes / data, extract the REUSABLE winning PATTERN(S) — capture the STYLE, never copy "
       "exact wording or facts. For EACH distinct post return a markdown block exactly like:\n"
       "- TOPIC: <short>\n  HOOK: \"<grabby opener style>\"\n  HEADLINE STYLE: \"<bold CAPS shape>\"\n"
       "  WHY IT WORKS: <one line>\nReturn 1-5 blocks. If the input is just an idea/note, still shape it "
       "into one block. Output ONLY the blocks, nothing else.")

def _api(method, **kw):
    return requests.post(f"https://api.telegram.org/bot{C.TG_TOKEN}/{method}", data=kw, timeout=60).json()

def _file(file_id):
    fp = _api("getFile", file_id=file_id)["result"]["file_path"]
    return requests.get(f"https://api.telegram.org/file/bot{C.TG_TOKEN}/{fp}", timeout=90).content

def _llm(content):
    client = OpenAI(api_key=C.OPENAI_API_KEY)
    r = client.chat.completions.create(model="gpt-4o-mini", temperature=0.2,
        messages=[{"role": "system", "content": SYS}, {"role": "user", "content": content}])
    return r.choices[0].message.content.strip()

def _from_image(b):
    b64 = "data:image/jpeg;base64," + base64.b64encode(b).decode()
    return _llm([{"type": "text", "text": "Extract the winning pattern(s) from this post screenshot:"},
                 {"type": "image_url", "image_url": {"url": b64, "detail": "low"}}])

def _from_text(t):
    t = (t or "").strip()
    return _llm("Extract the winning pattern(s) from this:\n\n" + t[:8000]) if t else None

def _pdf_text(b):
    try:
        from pypdf import PdfReader
        return "\n".join((p.extract_text() or "") for p in PdfReader(io.BytesIO(b)).pages)[:8000]
    except Exception:
        return ""

def _excel_text(b):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(b), read_only=True, data_only=True)
        rows = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                rows.append(" ".join(str(c) for c in row if c is not None))
        return "\n".join(rows)[:8000]
    except Exception:
        return ""

def _printable_ratio(t):
    if not t:
        return 0.0
    return sum(1 for c in t if c.isprintable() or c in "\n\t\r") / len(t)

def _apify_text(b, top=15):
    """Apify Instagram scrapers export a JSON array of post objects. Instead of feeding raw JSON,
    pull each post's caption + engagement, rank by likes+comments, and hand the LLM only the
    TOP performers' captions — so it learns the style of what actually works, not metadata noise.
    Returns None if this doesn't look like such an export (caller falls back to raw text)."""
    try:
        data = json.loads(b.decode("utf-8", "ignore"))
    except Exception:
        return None
    if isinstance(data, dict):                       # some exports wrap the list under a key
        for k in ("items", "data", "results", "posts"):
            if isinstance(data.get(k), list):
                data = data[k]; break
    if not isinstance(data, list) or not data or not isinstance(data[0], dict):
        return None
    posts = []
    for d in data:
        if not isinstance(d, dict):
            continue
        cap = d.get("caption") or d.get("text") or d.get("title") or ""
        if not isinstance(cap, str) or len(cap.strip()) < 15:
            continue
        eng = (d.get("likesCount") or d.get("likes") or 0) + (d.get("commentsCount") or d.get("comments") or 0)
        try:
            eng = int(eng)
        except Exception:
            eng = 0
        posts.append((eng, cap.strip()))
    if not posts:
        return None
    posts.sort(key=lambda x: x[0], reverse=True)
    lines = [f"[{e} engagement] {c[:400]}" for e, c in posts[:top]]
    return "TOP-PERFORMING posts (ranked by likes+comments) — learn the STYLE only:\n\n" + "\n\n".join(lines)

def _handle(msg):
    """Return extracted markdown block(s) for one message, or None if nothing usable.
    Accepts: image, plain text, and ANY file — code (.py/.java/.js…), data (.csv/.json),
    .txt/.md, PDF, Excel. Pure-binary files (zip/exe/video) can't be learned from -> None."""
    if msg.get("photo"):                                  # image
        return _from_image(_file(msg["photo"][-1]["file_id"]))
    doc = msg.get("document")
    if doc:                                               # any file
        b = _file(doc["file_id"])
        mime = (doc.get("mime_type") or "").lower()
        name = (doc.get("file_name") or "").lower()
        if mime.startswith("image/"):
            return _from_image(b)
        if mime == "application/pdf" or name.endswith(".pdf"):
            return _from_text(_pdf_text(b))
        if name.endswith((".xlsx", ".xlsm")) or "spreadsheet" in mime:
            return _from_text(_excel_text(b))
        if name.endswith(".json") or mime == "application/json":
            apify = _apify_text(b)                         # Apify IG export -> top-post captions
            if apify:
                return _from_text(apify)
        text = b.decode("utf-8", "ignore")                # code / data / text of any extension
        return _from_text(text) if _printable_ratio(text) > 0.85 else None
    txt = (msg.get("text") or msg.get("caption") or "").strip()   # plain text / image caption
    if txt and not txt.startswith("/"):                  # ignore bot commands like /start
        return _from_text(txt)
    return None

def run():
    if not C.TG_TOKEN:
        print("[ingest] no telegram token"); return
    offset = int(OFFSET_FILE.read_text().strip()) if OFFSET_FILE.exists() else 0
    updates = _api("getUpdates", offset=offset, allowed_updates=json.dumps(["message"])).get("result", [])
    added, last = 0, offset
    for u in updates:
        last = u["update_id"] + 1
        msg = u.get("message", {}); chat = (msg.get("chat") or {}).get("id")
        try:
            block = _handle(msg)
            if block:
                with open(EXAMPLES, "a", encoding="utf-8") as f:
                    f.write("\n\n" + block + "\n")
                added += 1
                if chat:
                    _api("sendMessage", chat_id=chat, text="✅ Learned from that — added to the library.")
        except Exception as e:
            print("[ingest] one failed:", e)
            if chat:
                _api("sendMessage", chat_id=chat, text="⚠️ Couldn't read that one, sorry — try another.")
    if last != offset:
        OFFSET_FILE.write_text(str(last))
    print(f"[ingest] {len(updates)} updates, {added} new examples added")

if __name__ == "__main__":
    run()
